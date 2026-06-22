import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="0F3D63", end_color="0F3D63", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 40)


def build_finance_excel(rentals_qs, expense_entries, owner_payouts, salary_payments, month, year) -> bytes:
    """
    Sheet 1: every rental order in the period - rental id, customer name/number,
             vehicle number, base amount, gst %, gst amount, total.
    Sheet 2: all expenses for the period (custom expense entries, owner
             payouts, staff salary payments) in one consolidated list.
    """
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Rental Income"
    headers1 = [
        "Rental ID", "Invoice No.", "Customer Name", "Customer Number",
        "Vehicle Number", "Base Amount", "GST %", "GST Amount", "Total Amount",
        "Amount Paid", "Balance Due", "Status", "Date",
    ]
    ws1.append(headers1)
    _style_header(ws1)

    for r in rentals_qs:
        ws1.append([
            r.id, r.invoice_number, r.customer.full_name, r.customer.phone,
            r.vehicle.registration_number, float(r.base_amount), float(r.gst_percent_snapshot),
            float(r.gst_amount), float(r.total_amount), float(r.amount_paid),
            float(r.balance_due), r.get_status_display(),
            r.created_at.strftime("%d-%m-%Y"),
        ])
    _autosize(ws1)

    ws2 = wb.create_sheet("Expenses")
    headers2 = ["Type", "Category / Recipient", "Title / Description", "Amount", "Date", "Notes"]
    ws2.append(headers2)
    _style_header(ws2)

    for e in expense_entries:
        ws2.append(["Custom Expense", e.get_category_display(), e.title, float(e.amount), e.date.strftime("%d-%m-%Y"), e.notes])

    for p in owner_payouts:
        ws2.append(["Owner Payout", p.owner.name, p.get_payout_type_display(), float(p.amount), p.paid_at.strftime("%d-%m-%Y"), p.notes])

    for s in salary_payments:
        ws2.append(["Staff Salary", s.staff.full_name, f"Salary {s.month}/{s.year}", float(s.final_amount), (s.paid_at.strftime("%d-%m-%Y") if s.paid_at else ""), s.notes])

    _autosize(ws2)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
